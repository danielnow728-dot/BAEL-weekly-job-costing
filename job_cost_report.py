import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import io

def compute_report(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, float]]:
    # --- filter BAEL jobs
    df = df[df["All Jobs"].astype(str).str.contains("BAEL", case=False, na=False)].copy()

    # --- numeric coercions
    # --- numeric coercions
    for c in ["Reg", "OT", "Reg.1", "PerDiem", "Travel"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
        else:
            df[c] = 0.0

    # --- remove non-data lines (keep if hour, perdiem, or travel)
    df = df[(df["Reg"].fillna(0) > 0) | (df["OT"].fillna(0) > 0) | (df["PerDiem"].fillna(0) > 0) | (df["Travel"].fillna(0) > 0)].copy()
    df["Job Name"] = df["All Jobs"].astype(str)

    # --- aggregate duplicates per employee per job
    agg = (
        df.groupby(["Job Name", "Employee Name"], as_index=False)
          .agg({"Reg": "sum", "OT": "sum", "Reg.1": "sum", "PerDiem": "sum", "Travel": "sum"})
    )

    # --- cost calculations
    agg["Regular Hourly Rate"] = agg["Reg.1"] / agg["Reg"].replace({0: np.nan})
    agg["Overtime Rate"] = agg["Regular Hourly Rate"] * 1.5
    agg["Loaded Regular Rate"] = agg["Regular Hourly Rate"] * 1.24
    agg["Loaded Overtime Rate"] = agg["Overtime Rate"] * 1.24

    agg["Loaded REG Cost"] = agg["Loaded Regular Rate"] * agg["Reg"]
    agg["Loaded OT Cost"] = agg["Loaded Overtime Rate"] * agg["OT"]
    agg["Total Loaded Cost"] = agg["Loaded REG Cost"].fillna(0) + agg["Loaded OT Cost"].fillna(0)
    both_cost_nan = agg["Loaded REG Cost"].isna() & agg["Loaded OT Cost"].isna()
    agg.loc[both_cost_nan, "Total Loaded Cost"] = np.nan

    # --- billing rate rules
    jn = agg["Job Name"].astype(str)
    is_lrs = jn.str.contains("LRS", case=False, na=False)
    is_avs_los = jn.str.contains("AVS", case=False, na=False) | jn.str.contains("LOS", case=False, na=False)

    reg_bill_rate = np.where(is_lrs, 48.0, np.where(is_avs_los, 46.0, np.nan))
    ot_bill_rate = np.where(is_lrs, 70.40, np.where(is_avs_los, 67.23, np.nan))

    agg["Regular Billing"] = agg["Reg"] * reg_bill_rate
    agg["Overtime Billing"] = agg["OT"] * ot_bill_rate
    agg["Total Billing"] = agg["Regular Billing"].fillna(0) + agg["Overtime Billing"].fillna(0)
    both_bill_nan = agg["Regular Billing"].isna() & agg["Overtime Billing"].isna()
    agg.loc[both_bill_nan, "Total Billing"] = np.nan

    # --- gross profit & margin
    agg["Gross Profit $"] = agg["Total Billing"] - agg["Total Loaded Cost"]
    agg["Gross Margin %"] = agg["Gross Profit $"] / agg["Total Billing"]
    agg.loc[agg["Total Billing"].isna() | (agg["Total Billing"] == 0), "Gross Margin %"] = np.nan
    agg.loc[agg["Total Billing"].isna() | agg["Total Loaded Cost"].isna(), "Gross Profit $"] = np.nan

    # --- weighted average base wage per job (base already excludes OT premium)
    # --- weighted average base wage per job (base already excludes OT premium)
    agg["Total Hours"] = agg["Reg"].fillna(0) + agg["OT"].fillna(0)
    job_avg = {}
    for job, g in agg.groupby("Job Name"):
        denom = g["Total Hours"].sum()
        if denom == 0:
            job_avg[job] = np.nan
        else:
            job_avg[job] = (g["Regular Hourly Rate"] * g["Total Hours"]).sum(skipna=True) / denom

    # Separate out PerDiem and Travel into job-level summaries
    # We remove them from individual rows to prevent double-counting
    job_expenses = (
        agg.groupby("Job Name", as_index=False)
           .agg({"PerDiem": "sum", "Travel": "sum"})
    )
    
    # Drop them from the employee rows so we don't accidentally print them on every employee line
    agg = agg.drop(columns=["PerDiem", "Travel"])
    
    # output sorting
    agg = agg.sort_values(["Job Name", "Employee Name"]).reset_index(drop=True)
    return agg, job_avg, job_expenses


def write_grouped_excel(agg: pd.DataFrame, job_avg: dict[str, float], job_expenses: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "BAEL Weekly Costs"

    # Styles (matches what we built together)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    main_title_font = Font(bold=True, size=16)
    subheader_fill = PatternFill("solid", fgColor="D9E1F2")
    info_fill = PatternFill("solid", fgColor="EDF2F9")
    bold_font = Font(bold=True)

    currency = "$#,##0.00"
    hours = "0.00"
    pct = "0.00%"

    thin = Side(style="thin", color="BFBFBF")
    thick = Side(style="thick", color="1F4E79")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = [
        "Employee Name",
        "Regular Hourly Rate",
        "Overtime Rate",
        "Loaded Regular Rate",
        "Loaded Overtime Rate",
        "REG Hours",
        "OT Hours",
        "Loaded REG Cost",
        "Loaded OT Cost",
        "Total Loaded Cost",
        "Regular Billing",
        "Overtime Billing",
        "Total Billing",
        "Gross Profit $",
        "Gross Margin %",
    ]

    # Column widths (tuned)
    widths = [28, 18, 14, 18, 18, 10, 10, 16, 16, 18, 16, 16, 16, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(cols))
    t = ws.cell(row=row, column=1, value="BAEL Jobs – Weekly Cost, Billing, Gross Profit & Avg Wage Summary")
    t.font = main_title_font
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 2

    for job, g in agg.groupby("Job Name", sort=True):
        # Job title row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(cols))
        jc = ws.cell(row=row, column=1, value=job)
        jc.font = title_font
        jc.fill = subheader_fill
        jc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        # Avg wage row split into label/value/text (as you requested)
        avg = job_avg.get(job, np.nan)

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        lbl = ws.cell(row=row, column=1, value="Weighted Avg Hourly Wage (Base)")
        lbl.fill = info_fill
        lbl.font = Font(italic=True, color="1F4E79")

        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        val = ws.cell(row=row, column=4, value=None if np.isnan(avg) else float(avg))
        val.number_format = currency
        val.font = Font(bold=True)
        val.fill = info_fill

        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=len(cols))
        txt = ws.cell(
            row=row,
            column=6,
            value="Weighted by (REG Hours + OT Hours); OT converted to base by ÷1.5",
        )
        txt.fill = info_fill
        txt.font = Font(italic=True, color="1F4E79")
        row += 1

        # Header row
        for c, name in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[row].height = 30
        row += 1

        # Data rows (values)
        for _, r in g.iterrows():
            vals = [
                r["Employee Name"],
                r["Regular Hourly Rate"],
                r["Overtime Rate"],
                r["Loaded Regular Rate"],
                r["Loaded Overtime Rate"],
                r["Reg"],
                r["OT"],
                r["Loaded REG Cost"],
                r["Loaded OT Cost"],
                r["Total Loaded Cost"],
                r["Regular Billing"],
                r["Overtime Billing"],
                r["Total Billing"],
                r["Gross Profit $"],
                r["Gross Margin %"],
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=None if (isinstance(v, float) and np.isnan(v)) else v)

                if c in [2, 3, 4, 5, 8, 9, 10, 11, 12, 13, 14]:
                    cell.number_format = currency
                elif c in [6, 7]:
                    cell.number_format = hours
                elif c == 15:
                    cell.number_format = pct

                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left" if c == 1 else "right", vertical="center")
            row += 1

        # Retrieve Job Expenses
        job_expense_row = job_expenses[job_expenses["Job Name"] == job]
        per_diem = float(job_expense_row["PerDiem"].iloc[0]) if not job_expense_row.empty else 0.0
        travel = float(job_expense_row["Travel"].iloc[0]) if not job_expense_row.empty else 0.0
        
        # Add PerDiem Row if > 0
        if per_diem > 0:
            ws.cell(row=row, column=1, value="PerDiem").font = bold_font
            perdiem_cost_cell = ws.cell(row=row, column=10, value=per_diem) # Total Loaded Cost
            perdiem_bill_cell = ws.cell(row=row, column=13, value=per_diem) # Total Billing
            perdiem_cost_cell.number_format = currency
            perdiem_bill_cell.number_format = currency
            for c in range(1, len(cols) + 1):
                ws.cell(row=row, column=c).border = thin_border
            row += 1
            
        # Add Travel Row if > 0
        if travel > 0:
            ws.cell(row=row, column=1, value="Travel").font = bold_font
            travel_cost_cell = ws.cell(row=row, column=10, value=travel) # Total Loaded Cost
            travel_bill_cell = ws.cell(row=row, column=13, value=travel) # Total Billing
            travel_cost_cell.number_format = currency
            travel_bill_cell.number_format = currency
            for c in range(1, len(cols) + 1):
                ws.cell(row=row, column=c).border = thin_border
            row += 1

        # JOB TOTAL row (no hourly rate totals)
        ws.cell(row=row, column=1, value="JOB TOTAL").font = bold_font

        reg_hours = float(g["Reg"].sum())
        ot_hours = float(g["OT"].sum())
        loaded_reg = float(g["Loaded REG Cost"].sum(skipna=True))
        loaded_ot = float(g["Loaded OT Cost"].sum(skipna=True))
        base_total_cost = float(g["Total Loaded Cost"].sum(skipna=True))
        
        bill_reg = float(g["Regular Billing"].sum(skipna=True))
        bill_ot = float(g["Overtime Billing"].sum(skipna=True))
        base_total_bill = float(g["Total Billing"].sum(skipna=True))
        
        # Incorporate pass-throughs into the aggregate totals
        total_cost = base_total_cost + per_diem + travel
        total_bill = base_total_bill + per_diem + travel

        gp = total_bill - total_cost
        gm = (gp / total_bill) if total_bill != 0 else np.nan

        totals_map = {
            6: (reg_hours, hours),
            7: (ot_hours, hours),
            8: (loaded_reg, currency),
            9: (loaded_ot, currency),
            10: (total_cost, currency),
            11: (bill_reg, currency),
            12: (bill_ot, currency),
            13: (total_bill, currency),
            14: (gp, currency),
            15: (gm, pct),
        }

        for col, (val, fmt) in totals_map.items():
            cell = ws.cell(row=row, column=col, value=None if (isinstance(val, float) and np.isnan(val)) else val)
            cell.number_format = fmt

        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = bold_font
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
            cell.border = Border(left=thin, right=thin, top=thick, bottom=thin)
            cell.alignment = Alignment(horizontal="left" if c == 1 else "right", vertical="center")

        row += 2

    # Save to a virtual file (io.BytesIO) to be returned directly by FastAPI
    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    return virtual_workbook

def update_running_master(agg: pd.DataFrame, job_expenses: pd.DataFrame, week_name: str, master_filepath: str):
    """
    Updates or creates a master running table tracking Total Billing and Margin 
    for each job across different weeks.
    
    Expected output structure matches the user's screenshot:
    Row 1 (Headers level 1): [Blank] | Week [name] | Week [name] | ...
    Row 2 (Headers level 2): [Blank] | Total Billing | Margin | Total Billing | Margin | ...
    Row 3+: Job Name | $val | %val | $val | %val | ...
    """
    import os
    from openpyxl import load_workbook
    
    # 1. Aggregate the new data per job
    # We need Total Billing and Gross Margin % per Job Name for this specific week.
    job_summary = {}
    for job, g in agg.groupby("Job Name"):
        # Base billing & cost without expenses
        base_bill = float(g["Total Billing"].sum(skipna=True))
        base_cost = float(g["Total Loaded Cost"].sum(skipna=True))
        
        # Pull expenses for this job
        job_expense_row = job_expenses[job_expenses["Job Name"] == job]
        per_diem = float(job_expense_row["PerDiem"].iloc[0]) if not job_expense_row.empty else 0.0
        travel = float(job_expense_row["Travel"].iloc[0]) if not job_expense_row.empty else 0.0
        
        # Add pass-throughs
        total_bill = base_bill + per_diem + travel
        total_cost = base_cost + per_diem + travel
        
        gp = total_bill - total_cost
        margin = (gp / total_bill) if total_bill != 0 else np.nan
        job_summary[job] = {"billing": total_bill, "margin": margin}
        
    # 2. Open or Create Master Workbook
    if os.path.exists(master_filepath):
        wb = load_workbook(master_filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Master Tracker"
        # Setup initial column A
        ws.column_dimensions['A'].width = 45
        
    # Find the next available column block (we need 2 columns per week)
    # Row 1 has the "Week ***" headers. Let's find the max column in row 1.
    max_col = ws.max_column
    if max_col == 1 and ws.cell(row=1, column=1).value is None:
        start_col = 2 # First time adding a week
    else:
        start_col = max_col + 1
        
    # Styles for Master Sheet
    header_fill = PatternFill("solid", fgColor="E2EFDA") # Light green header
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"), 
        right=Side(style="thin", color="BFBFBF"), 
        top=Side(style="thin", color="BFBFBF"), 
        bottom=Side(style="thin", color="BFBFBF")
    )
    
    # 3. Add Headers for the new week
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col+1)
    week_header = ws.cell(row=1, column=start_col, value=week_name)
    week_header.font = bold_font
    week_header.alignment = Alignment(horizontal="center")
    
    bill_header = ws.cell(row=2, column=start_col, value="Total Billing")
    margin_header = ws.cell(row=2, column=start_col+1, value="Margin")
    
    for cell in [week_header, bill_header, margin_header]:
        cell.fill = header_fill
        cell.border = thin_border
    
    # Adjust column widths
    ws.column_dimensions[get_column_letter(start_col)].width = 15
    ws.column_dimensions[get_column_letter(start_col+1)].width = 12
    
    # 4. Read existing jobs and their row numbers
    existing_jobs = {}
    # We assume jobs start at row 3, column 1
    max_row = ws.max_row
    for r in range(3, max_row + 1):
        job_name = ws.cell(row=r, column=1).value
        if job_name:
            existing_jobs[job_name] = r
            
    # 5. Insert data
    next_new_row = max_row + 1 if max_row >= 3 else 3
    
    currency_fmt = "$#,##0.00"
    pct_fmt = "0.00%"
    
    for job, data in job_summary.items():
        if job in existing_jobs:
            target_row = existing_jobs[job]
        else:
            # New job found, add it to the bottom
            target_row = next_new_row
            job_cell = ws.cell(row=target_row, column=1, value=job)
            job_cell.border = thin_border
            existing_jobs[job] = target_row
            next_new_row += 1
            
        # Write values
        bill_cell = ws.cell(row=target_row, column=start_col, value=data["billing"] if not np.isnan(data["billing"]) else None)
        margin_cell = ws.cell(row=target_row, column=start_col+1, value=data["margin"] if not np.isnan(data["margin"]) else None)
        
        bill_cell.number_format = currency_fmt
        margin_cell.number_format = pct_fmt
        
    # Save directly to the persistent disk path
    wb.save(master_filepath)

def check_if_week_exists(week_name: str, master_filepath: str) -> bool:
    """
    Checks if the given week_name is already present in the header row of the Master Tracker.
    """
    import os
    from openpyxl import load_workbook
    
    if not os.path.exists(master_filepath):
        return False
        
    try:
        wb = load_workbook(master_filepath, read_only=True)
        ws = wb.active
        # Check row 1 for the week_name
        for col in range(2, ws.max_column + 1):
            cell_val = ws.cell(row=1, column=col).value
            if cell_val and str(cell_val).strip() == str(week_name).strip():
                return True
    except Exception:
        pass
        
    return False
